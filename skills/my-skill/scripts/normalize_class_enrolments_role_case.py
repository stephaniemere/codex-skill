#!/usr/bin/env python3
"""Normalize Class Enrolments.role to lowercase (values-only safe).

This is used to ensure role values like "Student"/"Teacher" are consistently
lowercase ("student"/"teacher") in the final workbook.
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


def normalize_role_cell(value: object) -> object:
    if not isinstance(value, str):
        return value
    stripped = value.strip()
    if stripped == "":
        return ""
    return stripped.lower()


def main() -> int:
    parser = argparse.ArgumentParser(description="Lowercase Class Enrolments.role values")
    parser.add_argument("--input", required=True, help="Input .xlsx")
    parser.add_argument("--output", required=True, help="Output .xlsx")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)

    wb = load_workbook(input_path, data_only=False)

    sheet_name = None
    for candidate in ("Class Enrolments", "Class Enrollments"):
        if candidate in wb.sheetnames:
            sheet_name = candidate
            break
    if sheet_name is None:
        wb.close()
        raise SystemExit("Missing required sheet: Class Enrolments")

    ws = wb[sheet_name]

    # Determine header columns from row 1.
    headers = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is None:
            break
        headers.append(str(v).strip())
    col_by_header = {h: i + 1 for i, h in enumerate(headers)}

    role_col = col_by_header.get("role")
    class_id_col = col_by_header.get("classId")
    if role_col is None or class_id_col is None:
        wb.close()
        raise SystemExit("Missing required columns: classId and/or role")

    changed = 0
    row = 2
    while True:
        class_id = ws.cell(row=row, column=class_id_col).value
        if class_id is None or str(class_id).strip() == "":
            break
        cell = ws.cell(row=row, column=role_col)
        old = cell.value
        new = normalize_role_cell(old)
        if new != old:
            changed += 1
        cell.value = new
        row += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()

    print(f"Normalized {changed} role cells in '{sheet_name}'.")
    print(f"Wrote {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

